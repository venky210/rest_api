from django.shortcuts import render,get_object_or_404

# Create your views here.
from rest_framework.decorators import api_view,permission_classes,authentication_classes
from rest_framework.response import Response
from api_app.serializers import *
from django.contrib.auth.hashers import make_password
from django.contrib.auth import authenticate,login
from rest_framework_simplejwt.tokens import RefreshToken
from rest_framework import status
from rest_framework.views import APIView
from rest_framework.permissions import IsAuthenticated

from django.contrib.sessions.models import Session
from rest_framework.settings import api_settings
from openpyxl import load_workbook

# from rest_framework.authentication import TokenAuthentication

@api_view(['POST'])
def Registration(request):
    if request.method=='POST':
        serializer=RegisterSerializers(data=request.data)
        if serializer.is_valid():
            serializer.save()
            return Response(serializer.data)
        return Response(serializer.errors)
   
    


@api_view(['POST'])
def login(request):
    if request.method == 'POST':
        username = request.data.get('username')  
        password = request.data.get('password')
        if not username or not password:
            return Response({'error': 'Both username/email and password are required.'}, 
                            status=status.HTTP_400_BAD_REQUEST)
        user = authenticate(username=username, password=password)
        if user:
            refresh = RefreshToken.for_user(user)
            return Response({'refresh': str(refresh), 'access': str(refresh.access_token),'message':'User Login Successfully...'}, status=status.HTTP_200_OK)
       
        return Response({'error': 'Invalid username/email or password.'}, 
                        status=status.HTTP_401_UNAUTHORIZED)
    
    



@api_view(['POST'])
@permission_classes([IsAuthenticated])
def ChangePassword(request):
    if request.method=='POST':
        serializer = ChangePasswordSerializer(data=request.data)
        if serializer.is_valid():
            user = request.user
            old_password = serializer.validated_data['old_password']
            new_password = serializer.validated_data['new_password']
            
            if not user.check_password(old_password):
                return Response({'error': 'Old password is incorrect.'}, status=status.HTTP_400_BAD_REQUEST)

            # Set the new password
            user.set_password(new_password)
            user.save()
            return Response({'message': 'Password changed successfully.'}, status=status.HTTP_200_OK)
        else:
            return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)





@api_view(['PUT'])
@permission_classes([IsAuthenticated])
def editprofile(request):
    if request.method == 'PUT':
        user = request.user  
        serializer = editprofileserializer(instance=user, data=request.data)
        if serializer.is_valid():
            serializer.save()
            return Response({"data": serializer.data, "message": "Profile successfully changed"}, status=status.HTTP_200_OK)
        return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)
    

@api_view(['POST'])
@permission_classes([IsAuthenticated])
def CreateProduct(request):
    if request.method=='POST':
        if request.user.dealer:
            serializers=createproductserializer(data=request.data)
            if serializers.is_valid():
                serializers.save(dealer=request.user)
                return Response(serializers.data,status=status.HTTP_201_CREATED)
            return Response(serializers.errors,status=status.HTTP_400_BAD_REQUEST)
        return Response('Invaild Login...')



@api_view(['GET'])
@permission_classes([IsAuthenticated])
def dealer_product_list(request):
    if request.user.dealer:
        products = Product.objects.filter(dealer=request.user)
        serializer = createproductserializer(products, many=True)
        return Response(serializer.data)
    return Response('Invaild Login...')



@api_view(['GET'])
@permission_classes([IsAuthenticated])
def allproducts(request):
    if request.user.user:
        products=Product.objects.filter(status='approved')
        serializers=createproductserializer(products,many=True)
        return Response(serializers.data)
    
    return Response('Invaild Login...')



@api_view(['POST'])
@permission_classes([IsAuthenticated])
def updateproduct(request, product_id):
    if request.user.dealer:
        product = get_object_or_404(Product, pk=product_id, dealer=request.user)
        if request.method == 'POST':
            serializer = productupdateserializer(product, data=request.data)
            if serializer.is_valid():
                serializer.save()
                return Response(serializer.data)
            return Response(serializer.errors)  


@api_view(['DELETE'])
@permission_classes([IsAuthenticated])
def productdelete(request,product_id):
    if request.user.dealer:
        product=get_object_or_404(Product,pk=product_id,dealer=request.user)
        if request.method=='DELETE':  
            product.delete()
            return Response('Product Delete Successfully...')   
      




@api_view(['POST'])
@permission_classes([IsAuthenticated])
def addwishlist(request,product_id):
    if request.user.user:
        product = get_object_or_404(Product, pk=product_id)
        existing_wishlist_item = Wishlist.objects.filter(user=request.user, product=product).first()

        if existing_wishlist_item:
            return Response({"detail": "Product already in wishlist."}, status=status.HTTP_400_BAD_REQUEST)

        serializer = wishlistserializer(data={"product": product.id}, context={"request": request})
        if serializer.is_valid():
            serializer.save(user=request.user, product=product)
            return Response({"data":serializer.data,"message":'Product Addwishlist Successfully...' },status=status.HTTP_201_CREATED)
        return Response(serializer.errors,status=status.HTTP_400_BAD_REQUEST)




@api_view(['GET'])
@permission_classes([IsAuthenticated])
def wishlist(request):
    if request.user.user:
        wishlist_items = Wishlist.objects.filter(user=request.user)
        serializer = wishlistserializer(wishlist_items, many=True)
        return Response(serializer.data, status=status.HTTP_200_OK)




# @api_view(['DELETE'])
# @permission_classes([IsAuthenticated])
# def removewishlist(request,wishlist_id):
#     product=get_object_or_404(Wishlist,pk=wishlist_id,user=request.user)
#     if request.method=='DELETE':
#         product.delete()
#         return Response('wishlist Product delete successfully...')
   


@api_view(['POST'])
@permission_classes([IsAuthenticated])
def removewishlist(request, product_id):
    wishlist_item = Wishlist.objects.filter(user=request.user).first()  
    if not wishlist_item:
        return Response({"message": "Product is not in your wishlist."}, status=status.HTTP_400_BAD_REQUEST)  
    wishlist_item.delete()  
    return Response({"message": "Product removed from wishlist successfully."}, status=status.HTTP_200_OK)


# @api_view(['POST'])
# @permission_classes([IsAuthenticated])
# def addtocart(request,product_id):
#     product=get_object_or_404(Product,product_id)
#     if request.method=='POST':
#         serializers=addcartserializer(data=request.data,many=True)
#         if serializers.is_valid():
#             serializers.save(user=request.user,product=product)
#             return Response(serializers.data,'Add To cart successfully...')
#         return Response(serializers.error_messages)





# @api_view(['POST'])
# @permission_classes([IsAuthenticated])
# def addtocart(request,product_id):
#     if request.user.user:
#         product = get_object_or_404(Product, pk=product_id)
       
#         serializer = addcartserializer(data=request.data)
#         if serializer.is_valid():
#             serializer.save(user=request.user, product=product)
#             return Response(serializer.data,'Addwishlist Successfully...' ,status=status.HTTP_201_CREATED)
#         return Response(serializer.errors,status=status.HTTP_400_BAD_REQUEST)



@api_view(['POST'])
@permission_classes([IsAuthenticated])
def addtocart(request, product_id):
    product = get_object_or_404(Product, pk=product_id)
    existing_cart_item = Cart.objects.filter(user=request.user, product=product).first()

    if existing_cart_item:
        return Response({"detail": "Product already in cart."}, status=status.HTTP_400_BAD_REQUEST)

    serializer = Addcartserializer(data={"product": product.id}, context={"request": request})
    if serializer.is_valid():
        serializer.save(user=request.user, product=product)
        return Response({"data":serializer.data,"message": "Product added to cart successfully."}, status=status.HTTP_201_CREATED)
    return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)



@api_view(['GET'])
@permission_classes([IsAuthenticated])
def cartlist(request):
    if request.method == 'GET':
        if request.user.user:
            cart_items = Cart.objects.filter(user=request.user)
            serializer = Addcartserializer(cart_items, many=True)
            return Response(serializer.data, status=status.HTTP_200_OK)


@api_view(['POST'])
@permission_classes([IsAuthenticated])
def removecart(request,product_id):
    cart=Cart.objects.filter(user=request.user).first()
    if not cart:
        return Response('Product Not in Cart...')
    cart.delete()
    return Response('Product Remove in cart successfully...')



@api_view(['POST'])
@permission_classes([IsAuthenticated])
def createcategory(request):
    if request.method=='POST':
        if request.user.admin:
            serializers=createcategoryserializer(data=request.data)
            if serializers.is_valid():
                serializers.save()
                return Response(serializers.data,status=status.HTTP_200_OK,)
            return Response(serializers.error_messages)
        return Response('Invaild Login...')
    

@api_view(['GET'])
@permission_classes([IsAuthenticated])
def categorylist(request):
    if request.method == 'GET':
        category_list=Category.objects.all()
        serializer=createcategoryserializer(category_list,many=True)
        return Response(serializer.data)
    

@api_view(['DELETE'])
@permission_classes([IsAuthenticated])
def deletecategory(request,category_id):
    category=get_object_or_404(Category,pk=category_id)
    if request.method == 'DELETE':
        if request.user.admin:
            category.delete()
            return Response('Category delete successfully...')
        


@api_view(['POST'])
@permission_classes([IsAuthenticated])
def updatecategory(request,category_id):
    if request.user.admin:
        category=get_object_or_404(Category,pk=category_id)
        if request.method=='POST':
            serializer=createcategoryserializer(category,data=request.data)
            if serializer.is_valid():
                serializer.save()
                return Response(serializer.data)
            return Response(serializer.error_messages)


@api_view(['GET'])
def categorydetails(request, category_id):
    category = Category.objects.filter(pk=category_id).first()
    products = Product.objects.filter(category=category)   
    serializer =createcategoryserializer(category)
    product_data = [{'name': product.pname, 'price': product.price, 'status':product.status} for product in products]
    
    return Response({
        'category': serializer.data,
        'products': product_data
    }, status=status.HTTP_200_OK)   




@api_view(['POST'])
@permission_classes([IsAuthenticated])
def updatestatus(request, product_id):
    if request.method == 'POST':
        if request.user.admin:
            try:
                product_obj = Product.objects.get(pk=product_id)
            except Product.DoesNotExist:
                return Response({"message": "Product does not exist"}, status=status.HTTP_404_NOT_FOUND)
            
            new_status = request.data.get('status')
            if new_status is not None:
                product_obj.status = new_status
                product_obj.save()
                return Response({"message": "Product status updated successfully"})
            else:
                return Response({"message": "Status field is required"}, status=status.HTTP_400_BAD_REQUEST)
        else:
            return Response('Invaild Login...')
    else:
        return Response({"message": "Method not allowed"}, status=status.HTTP_405_METHOD_NOT_ALLOWED)




@api_view(['GET'])
@permission_classes([IsAuthenticated])
def lowtohigh(request):
    product=Product.objects.order_by('price')
    serializer=createproductserializer(product,many=True)
    return Response(serializer.data)


@api_view(['GET'])
@permission_classes([IsAuthenticated])
def hightolow(request):
    product=Product.objects.order_by('-price')
    serializer=createproductserializer(product,many=True)
    return Response(serializer.data)




@api_view(['POST'])
@permission_classes([IsAuthenticated])
def import_from_excel(request):
    if request.method == 'POST' and 'excel_file' in request.FILES:
        excel_file = request.FILES['excel_file']
        try:
            
            wb = load_workbook(excel_file)
            ws = wb.active

            for row in ws.iter_rows(min_row=2, values_only=True):
                pname, price, img, category_name = row
                
                
                category_instance, _ = Category.objects.get_or_create(category=category_name)

               
                product_data = {
                    'pname': pname,
                    'price': price,
                    'img': img,
                    'category': category_instance.id  
                }
                
                
                serializer =createproductserializer(data=product_data)
                if serializer.is_valid():
                    
                    serializer.save(dealer=request.user)
                else:
                   
                    return Response(serializer.errors, status=400)
            
            return Response({'message': 'Products imported successfully'}, status=201)
        except Exception as e:
           
            return Response({'error_message': str(e)}, status=500)
    else:
        return Response({'error_message': 'No Excel file provided'}, status=400)
    


@api_view(['POST'])
@permission_classes([IsAuthenticated])
def rating(request, product_id):
    try:
        product = Product.objects.get(pk=product_id)
    except Product.DoesNotExist:
        return Response({"error": "Product does not exist"}, status=404)

    if request.method == 'POST':
        if request.user.user:
            serializer = ratingserializer(data=request.data)
            if serializer.is_valid():
                serializer.save(product=product, user=request.user)
                return Response(serializer.data, status=201)
            return Response(serializer.errors, status=400)
        
@api_view(['GET'])
@permission_classes([IsAuthenticated])
def ratinglist(request):
    if request.user.dealer:
        ratings = Rating.objects.all()
        serializer =ratingserializer(ratings, many=True)
        return Response(serializer.data)
    else:
        return Response({'detail': 'Only dealers can access this endpoint'}, status=status.HTTP_403_FORBIDDEN)


@api_view(['POST'])
@permission_classes([IsAuthenticated])
def createcoupon(request):
    if request.method == 'POST':
            if request.user.dealer:
                serializer = couponserializer(data=request.data)
                if serializer.is_valid():
                    serializer.save()
                    return Response(serializer.data, status=status.HTTP_201_CREATED)
                return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)



@api_view(['GET'])
@permission_classes([IsAuthenticated])
def coupon_list(request):
    if request.method == 'GET':
        coupons = Coupon.objects.all()
        serializer = couponserializer(coupons, many=True)
        return Response(serializer.data)

  
@api_view(['PUT', 'DELETE'])
@permission_classes([IsAuthenticated])
def coupon_detail(request, pk):
    if request.user.dealer:
        try:
            coupon = Coupon.objects.get(pk=pk)
        except Coupon.DoesNotExist:
            return Response(status=status.HTTP_404_NOT_FOUND)
        if request.method == 'PUT':
            serializer = couponserializer(coupon, data=request.data)
            if serializer.is_valid():
                serializer.save()
                return Response({"data":serializer.data,"message":"Coupon Upadate Successfully..."})
            return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)

        elif request.method == 'DELETE':
            coupon.delete()
            return Response("Coupon Delete Successfully...",status=status.HTTP_204_NO_CONTENT)
        


@api_view(['POST'])
@permission_classes([IsAuthenticated])
def address(request):
    if request.method == 'POST':
        if request.user.user:
            serializer = useraddressserializers(data=request.data)
            if serializer.is_valid():
                serializer.save()
                return Response(serializer.data, status=status.HTTP_201_CREATED)
            else:
                return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)
        else:
            return Response("User not authorized to create address", status=status.HTTP_403_FORBIDDEN)
        
@api_view(['GET'])
@permission_classes([IsAuthenticated])
def address_list(request):
    if request.method=='GET':
        if request.user.user:
            address=UserAddress.objects.all()
            serializer=useraddressserializers(address,many=True)
            return Response(serializer.data)
        else:
            return Response(serializer.error)
        
@api_view(['PUT','DELETE'])
@permission_classes([IsAuthenticated])
def address_update(request,pk):
    if request.user.user:
        try:
            address= UserAddress.objects.get(pk=pk)
        except UserAddress.DoesNotExist:
            return Response(status=status.HTTP_404_NOT_FOUND)
        if request.method=='PUT':
            serializer=useraddressserializers(address,data=request.data)
            if serializer.is_valid():
                serializer.save()
                return Response({"data":serializer.data,"message":"Address Update Successfully..."})
            else:
                return Response(serializer.errors)
            
        elif request.method=='DELETE':
            address.delete()
            return Response("Address Delete Successfully...")
            

@api_view(['GET', 'PUT'])
@permission_classes([IsAuthenticated])
def order(request):
    if request.method == 'GET':
        try:
            cart = Cart.objects.filter(user=request.user)
        except Cart.DoesNotExist:
            return Response('Cart does not exist...', status=status.HTTP_404_NOT_FOUND)
        
        cart_items_data = []
        total_amount = 0
        
        for cart_item in cart:
            item_data = {
                'product_name': cart_item.product.pname,
                'price': cart_item.product.price,
                'quantity': cart_item.quantity,
                'image': cart_item.product.img.url if cart_item.product.img else None,
                'total_price': cart_item.product.price * cart_item.quantity
            }
            total_amount += item_data['total_price']
            cart_items_data.append(item_data)
        
       
        try:
            default_address = UserAddress.objects.get(user=request.user, is_default=True)
            shipping_address_serializer =useraddressserializers(default_address)
        except UserAddress.DoesNotExist:
            shipping_address_serializer = None
        
        shipping_address_data = shipping_address_serializer.data if shipping_address_serializer else {}

        return Response({"cart_items": cart_items_data, "total_amount": total_amount, "shipping_address": shipping_address_data,"payment":"Continue to Payment..."})
    
    elif request.method == 'PUT':
       
        try:
            default_address = UserAddress.objects.get(user=request.user, is_default=True)
        except UserAddress.DoesNotExist:
            return Response('Default shipping address does not exist', status=status.HTTP_404_NOT_FOUND)
        
        address_serializer =useraddressserializers(default_address, data=request.data, partial=True)
        if address_serializer.is_valid():
            address_serializer.save()
            return Response(address_serializer.data, status=status.HTTP_200_OK)
        return Response(address_serializer.errors, status=status.HTTP_400_BAD_REQUEST,)


@api_view(['GET'])
@permission_classes([IsAuthenticated])
def payment(request):
    if request.method=="GET":
        serializer=Payment.objects.all()
        return Response({"data":serializer,"message":"Payment List..."})



